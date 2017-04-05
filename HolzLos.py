# excel
from openpyxl import *
from openpyxl.cell import get_column_letter
# ui
import tkinter as tk
from tkinter.filedialog import askopenfilename, asksaveasfilename
from tkinter.messagebox import showerror, showinfo
# util
import heapq
import random
import bisect

class ConfigWindow:
    def __init__(self):
        # setup ui
        self.root = tk.Tk()
        self.root.title = "HolzLos-Configurator"
        losn_lbl = tk.Label(self.root, text="Anzahl der Lose")
        losn_lbl.pack()
        losn_box = tk.Text(self.root, height=1, width=30)
        losn_box.pack()
        losn_box.insert(tk.END, "20")
        self.losn_box = losn_box

        listsize_lbl = tk.Label(self.root, text="Mindest-Listengröße der Lose")
        listsize_lbl.pack()
        listsize_box = tk.Text(self.root, height=1, width=30)
        listsize_box.pack()
        listsize_box.insert(tk.END, "20")
        self.listsize_box = listsize_box
        
        opt_steps_lbl = tk.Label(self.root, text="Anzahl der Optimierungsschritte")
        opt_steps_lbl.pack()
        opt_steps_box = tk.Text(self.root, height=1, width=30)
        opt_steps_box.pack()
        opt_steps_box.insert(tk.END, "1000")
        self.opt_steps_box = opt_steps_box
        
        self.button_text = "Starten"
        button = tk.Button(self.root, text=self.button_text, width=30, command=self.select_files)
        button.pack()
        self.button = button
        tk.mainloop()

    
    def select_files(self):
        try:
            losn = int(self.losn_box.get("1.0", tk.END))
            listsize = int(self.listsize_box.get("1.0", tk.END))
            opt_steps = int(self.opt_steps_box.get("1.0", tk.END))
        except ValueError:
            showerror("Eingabefehler", "Losanzahl, Listengröße und Optimierungsschritte müssen ganze Zahlen sein")
            return
        
        self.button['text'] = "Verarbeiten..."
        
        infile = askopenfilename(filetypes=(("Excel Dateien", "*.xlsx"),
                                 ("All files", "*.*")))
        if not infile:
            return

        outfile = asksaveasfilename(filetypes=(("Excel Dateien", "*.xlsx"),
                                    ("All files", "*.*") ),
                                    initialfile="Lose.xlsx")
        if not outfile:
            return
         
        try:
            t = TableDivider()
            t.read_file(infile)
            t.divide(losn, opt_steps)
            t.write_file(outfile, listsize)
        except TableDividerError as e:
            showerror("Fehler", "Fehler beim verarbeiten den Datei %s:\n %s" % (infile, e))
            return
        showinfo("Erfolg", "Ausgabedatei erfolgreich erzeugt")
        self.button['text'] = self.button_text

class TableDividerError(Exception):
    def __init__(self, value):
        self.value = value
    def __str__(self):
        return repr(self.value)

class Polter:
    def __init__(self, nr, ort, menge):
        self.nr = nr
        self.ort = ort
        self.menge = menge

    def __lt__(self, other):
        return self.menge < other.menge

class Los:
    def __init__(self):
        self.polter_list = list()
        self.summe = 0

    def add_polter(self, polter):
        self.summe += polter.menge
        self.polter_list.append(polter)

    def __lt__(self, other):
        return self.summe < other.summe

class TableDivider:
    def __init__(self):
        self.hhsums = None
        self.whsums = None
        self.harth_l = None
        self.weichh_l = None
        self.harth_lose = None
        self.weichh_lose = None
        self.divisions = 1
        
    def read_file(self, infile):
        try:
            wb = load_workbook(filename=infile)
        except:
            raise TableDividerError("Konnte Datei nicht öffnen")
        
        wsname = 'Kontrollliste'
        try:
            ws = wb[wsname]
        except:
            raise TableDividerError("Tabelle '%s' fehlt" % wsname)
        a = ""
        firstrow = 0
        while(not (a is "1" or a == 1)):
            firstrow = firstrow + 1
            a = ws.cell(row=firstrow, column=1).value

        self.harth_l = list()
        self.weichh_l = list()
        i = firstrow
        while(True):
            polternr = ws.cell(row=i, column=1).value
            if polternr is None:
                break
            wort = ws.cell(row=i, column=2).value
            harth = ws.cell(row=i, column=4).value
            weichh = ws.cell(row=i, column=5).value
            try:
                if harth:
                    self.harth_l.append(Polter(int(polternr), wort, float(harth)))
                elif weichh:
                    self.weichh_l.append(Polter(int(polternr), wort, float(weichh)))
                else:
                    raise TableDividerError("Holztyp von Polter %s unbekannt" % polternr)
                i += 1
            except ValueError:
                raise TableDividerError("Ungültiger Datentyp bei Los %s" % polternr)

    def divide(self, divisions, opt_steps):
        if not self.harth_l or not self.weichh_l:
            raise TableDividerError("data must be read first")
        self.divisions = divisions
        self.harth_l.sort()
        self.weichh_l.sort()

        self.harth_lose = self.divide_polter(self.harth_l)
        self.weichh_lose = self.divide_polter(self.weichh_l)

        # optimieren der aufteilung
        self.optimize(self.harth_lose, opt_steps)
        self.optimize(self.weichh_lose, opt_steps)

    def write_file(self, outfile, listsize):
        if not self.harth_lose or not self.weichh_lose:
            raise TableDividerError("data must be divided first")
        self.harth_lose.sort()
        self.weichh_lose.sort(reverse=True)

        # check if listsize is sufficient
        for n in range(self.divisions):
            listsize = max(listsize, len(self.harth_lose[n].polter_list) +
                           len(self.weichh_lose[n].polter_list))

        wb = Workbook()
        ws = wb.active
        ws.title = "Aufteilung"

        self.write_lose(ws, listsize)

        try:
            wb.save(outfile)
        except Exception:
            raise TableDividerError("Augabedatei '%s' konnte nicht gespeichert werden" % outfile)

    def divide_polter(self, polter_list):
        los_list = list(Los() for _ in range(self.divisions))
        while len(polter_list) > 0:
            los = heapq.heappop(los_list)
            polter = polter_list.pop()
            los.add_polter(polter)
            heapq.heappush(los_list, los)
        return los_list

    def optimize(self, holz_lose, opt_steps):
        steps = opt_steps
        holz_lose.sort()
        while steps > 0:
            maxLos = holz_lose[-1]
            minLos = holz_lose[0]
            losDiff = maxLos.summe - minLos.summe
            losSum = maxLos.summe + minLos.summe
            while steps > 0:
                steps -= 1
                combined = maxLos.polter_list + minLos.polter_list
                random.shuffle(combined)
                los1 = Los()
                diff = losSum
                while abs(diff - 2*combined[-1].menge) < abs(diff):
                    diff -= 2*combined[-1].menge
                    los1.add_polter(combined.pop())
                    #print("len: %i, los1.summe: %d, losTarget: %d" % (len(combined), los1.summe, losSum/2))
                    #print("dist1: %d, dist2: %d" % ((diff - 2*combined[-1].menge), diff))
                if abs(losSum - 2*los1.summe) < losDiff:
                    del(holz_lose[0])
                    del(holz_lose[-1])
                    los2 = Los()
                    for polter in combined:
                        los2.add_polter(polter)
                    self.bisect_insert(holz_lose, los1)
                    self.bisect_insert(holz_lose, los2)
                    break

    def number_dist(self, a, b):
        return abs(a - b)

    def bisect_insert(self, lst, x):
        lst.insert(bisect.bisect(lst, x), x)

    def write_lose(self, ws, listsize):
        height = listsize + 3
        indent = 1
        row = 3
        # define list of sum-cells for verification
        self.hhsums = list()
        self.whsums = list()
        # iterate for each los
        for losnr in range(1, len(self.harth_lose) + 1):
            self.write_losform(ws, row, indent, losnr, height)

            list_row = self.write_los(ws, row+2, indent, self.harth_lose[losnr - 1], 2)
            list_row = self.write_los(ws, list_row, indent, self.weichh_lose[losnr - 1], 3)
            if list_row > row + height - 1:
                raise TableDividerError("Nicht genug Zeilen pro los.")
            # increment row + indent
            if indent > 1:
                indent = 1
                row += height
            else:
                indent = 7
        # print verification sums
        hhstring = ",".join(self.hhsums)
        whstring = ",".join(self.whsums)
        ws.cell("M4").value = "Summe:"
        ws.cell("N4").value = "Durchschnitt"
        ws.cell("O4").value = "Min"
        ws.cell("P4").value = "Max"
        ws.cell("L5").value = "Hartholz:"
        ws.cell("M5").value = "=SUM(%s)" % hhstring
        ws.cell("N5").value = "=M5/%i" % self.divisions
        ws.cell("O5").value = "=MIN(%s)" % hhstring
        ws.cell("P5").value = "=MAX(%s)" % hhstring
        ws.cell("L6").value = "Weichholz: "
        ws.cell("M6").value = "=SUM(%s)" % whstring
        ws.cell("N6").value = "=M6/%i" % self.divisions
        ws.cell("O6").value = "=MIN(%s)" % whstring
        ws.cell("P6").value = "=MAX(%s)" % whstring

    def write_losform(self, ws, row, indent, losnr, height):
        ws.merge_cells(start_row=row, start_column=indent, end_row=row, end_column=indent+3)
        ws.cell(row=row, column=indent).value = "Losnummer: %i" % losnr
        ws.cell(row=row + 1, column=indent).value = "PolterNr."
        ws.cell(row=row + 1, column=indent+1).value = "Waldort"
        ws.cell(row=row + 1, column=indent+2).value = "Hartholz"
        ws.cell(row=row + 1, column=indent+3).value = "Weichholz"
        lastrow = row + height - 1
        ws.cell(row=lastrow, column=indent).value = "Summe:"
        hhsumcell = ws.cell(row=lastrow, column=indent + 2)
        hhsumcell.value = self.sum_formula(row + 2, lastrow - 1, indent + 2)
        whsumcell = ws.cell(row=lastrow, column=indent + 3)
        whsumcell.value = self.sum_formula(row + 2, lastrow - 1, indent + 3)

        # store the sum-cells for later verification
        
        self.whsums.append(whsumcell.coordinate)
        self.hhsums.append(hhsumcell.coordinate)

    def sum_formula(self, row_start, row_end, column):
        cletter = get_column_letter(column)
        return "=SUM(%s%i:%s%i)" % (cletter, row_start, cletter, row_end)

    def write_los(self, ws, row, indent, los, hoffset):
        for polter in los.polter_list:
            ws.cell(row=row, column=indent).value = polter.nr
            ws.cell(row=row, column=indent+1).value = polter.ort
            ws.cell(row=row, column=indent+hoffset).value = polter.menge
            row += 1
        return row

    
if __name__ == "__main__":
    ConfigWindow()
